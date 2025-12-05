import openpyxl
import json

# Ler o arquivo Excel
excel_path = "agentes.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Extrair cabeçalho (nomes dos municípios)
header = []
for cell in ws[1]:
    if cell.value:
        header.append(cell.value)

# Extrair dados (nomes das pessoas)
municipios_agentes = {}

for col_idx, municipio in enumerate(header, 1):
    agentes = []
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value:
            agentes.append(cell_value)
    
    if agentes:
        municipios_agentes[municipio] = agentes

# Salvar como JSON
with open('municipios_agentes.json', 'w', encoding='utf-8') as f:
    json.dump(municipios_agentes, f, ensure_ascii=False, indent=2)

print("✅ Arquivo municipios_agentes.json criado!")
print(json.dumps(municipios_agentes, ensure_ascii=False, indent=2))
